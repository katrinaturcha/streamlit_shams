from __future__ import annotations

import io
import re
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd
import streamlit as st
from openpyxl.styles import PatternFill

from header_log import build_header_change_log_from_bytes
from shams_parser import parse_all_sheets_from_bytes
from compare import compare_shams, comparison_stats
from utils import normalize_text_for_compare
from DB import DB_COLUMNS


# ================== STAGES ==================
STAGE_UPLOAD = "upload"
STAGE_SELECT_HEADERS = "select_headers"
STAGE_MAPPING = "mapping"
STAGE_HIERARCHY = "hierarchy"
STAGE_COMPARE = "compare"
STAGE_DB_MAPPING = "db_mapping"
STAGE_DB_EXPORT = "db_export"


# ================== CONFIG ==================
st.set_page_config(layout="wide")

BASE_DIR = Path(__file__).resolve().parent
SHAMS_PATH = BASE_DIR / "shams.xlsx"
DB_PATH = BASE_DIR / "shams_edit1.xlsx"

if not DB_PATH.exists():
    st.error("Файл shams_edit1.xlsx (имитация БД) не найден")
    st.stop()

if not SHAMS_PATH.exists():
    st.error("Файл shams.xlsx не найден")
    st.stop()


# ================== SESSION STATE ==================
def init_state():
    defaults = {
        # bytes
        "shams_bytes": None,
        "shams2_bytes": None,

        # headers
        "headers_old": None,
        "headers_new": None,
        "headers_new_selected": None,

        # mapping shams2 -> shams
        "column_mapping": None,

        # hierarchy step
        "hier_levels": 1,  # 1/2/3
        "hier_col_roles": None,  # {col: "L1"/"L2"/"L3"/"SKIP"/"COMMON"}

        # compare
        "df_compare": None,
        "compare_stats": None,

        # db mapping
        "db_column_mapping": None,  # {src_col: db_col}
        "db_mapping_saved": False,
        "db_cols_order": None,  # list of src cols in UI order (for export pairing)

        # sheets comparisons
        "df_groups_cmp": None,
        "df_classes_cmp": None,
        "df_subclasses_cmp": None,

        # stage
        "stage": STAGE_UPLOAD,
        "provider_has_groups": False,
        "provider_has_classes": False,
        # --- DB preload UI (draft) ---
        "selected_provider": "SHAMS",
        "pre_db_mapping": {
            "activity_code": None,
            "official_title_en": None,
            "authority": None,
            "service": None,
        },
        "manager_processed_bytes": None,
        "manager_db_mapping_result": None,

    }
    for k, v in defaults.items():
        st.session_state.setdefault(k, v)


init_state()


# ================== HELPERS ==================
def load_shams():
    if st.session_state.shams_bytes is None:
        with open(SHAMS_PATH, "rb") as f:
            st.session_state.shams_bytes = f.read()


def _normalize_code(val):
    if pd.isna(val):
        return None
    s = re.sub(r"[^0-9]", "", str(val))
    if len(s) < 5:
        return None
    return f"{s[:4]}.{s[4:].ljust(2,'0')[:2]}"


def load_db_df() -> pd.DataFrame:
    """Читает shams_edit1.xlsx и гарантирует наличие Subclass_code."""
    df_db = pd.read_excel(DB_PATH)

    if "Subclass_code" in df_db.columns:
        df_db["Subclass_code"] = df_db["Subclass_code"].apply(_normalize_code)
        return df_db

    if "Введите код бизнес-деятельности" in df_db.columns:
        df_db["Subclass_code"] = df_db["Введите код бизнес-деятельности"].apply(_normalize_code)
        return df_db

    if "Subclass" in df_db.columns:
        df_db["Subclass_code"] = df_db["Subclass"].apply(_normalize_code)
        return df_db

    raise ValueError(
        "В shams_edit1.xlsx не найден ключевой столбец "
        "(Subclass_code / Subclass / 'Введите код бизнес-деятельности')"
    )


def _norm_col(x: str) -> str:
    if x is None:
        return ""
    s = str(x).replace("\u00A0", " ")
    s = " ".join(s.split())
    return s.strip().lower()


def compare_level_like_subclass(
    df_old: pd.DataFrame,
    df_new: pd.DataFrame,
    key_col: str,
    desc_col: str,
    out_key_name: Optional[str] = None,
    out_desc_name: str = "Description",
) -> pd.DataFrame:
    """
    Сравнение уровня (Group/Class/Subclass) по ключу и описанию:
    выход: status, <key>, Description (в виде OLD/NEW как в Subclass)
    """
    out_key_name = out_key_name or key_col

    if df_old is None or df_old.empty:
        df_old = pd.DataFrame(columns=[key_col, desc_col])
    if df_new is None or df_new.empty:
        df_new = pd.DataFrame(columns=[key_col, desc_col])

    df_old = df_old.copy()
    df_new = df_new.copy()

    df_old.columns = [str(c).strip() for c in df_old.columns]
    df_new.columns = [str(c).strip() for c in df_new.columns]

    if key_col not in df_old.columns and key_col not in df_new.columns:
        return pd.DataFrame(columns=["status", out_key_name, out_desc_name])

    need_old = [c for c in [key_col, desc_col] if c in df_old.columns]
    need_new = [c for c in [key_col, desc_col] if c in df_new.columns]
    df_old = df_old[need_old].copy()
    df_new = df_new[need_new].copy()

    df_old = df_old.add_suffix("_old").rename(columns={f"{key_col}_old": key_col})
    df_new = df_new.add_suffix("_new").rename(columns={f"{key_col}_new": key_col})

    merged = pd.merge(df_old, df_new, on=key_col, how="outer", indicator=True)

    def _initial_status(m):
        if m == "left_only":
            return "deleted"
        if m == "right_only":
            return "added"
        return "potentially_changed"

    merged["status"] = merged["_merge"].apply(_initial_status)

    old_desc = f"{desc_col}_old"
    new_desc = f"{desc_col}_new"

    def _final_status(row):
        stt = row["status"]
        if stt in ("added", "deleted"):
            return stt
        old_v = normalize_text_for_compare(row.get(old_desc, ""))
        new_v = normalize_text_for_compare(row.get(new_desc, ""))
        return "changed" if old_v != new_v else "not changed"

    merged["status"] = merged.apply(_final_status, axis=1)

    def _fmt(row):
        stt = row["status"]
        o = "" if pd.isna(row.get(old_desc)) else str(row.get(old_desc)).strip()
        n = "" if pd.isna(row.get(new_desc)) else str(row.get(new_desc)).strip()
        if stt == "changed":
            return f"OLD: {o}\nNEW: {n}".strip()
        if stt == "deleted":
            return o.strip() if o else ""
        if stt == "added":
            return n.strip() if n else ""
        return ""

    merged[out_desc_name] = merged.apply(_fmt, axis=1)

    out = merged[[key_col, "status", out_desc_name]].rename(columns={key_col: out_key_name})
    out = out[["status", out_key_name, out_desc_name]]
    return out


def _build_export_df(
    df_compare: pd.DataFrame,
    db_df: pd.DataFrame,
    db_map: dict,
    cols_order: list[str],
) -> pd.DataFrame:
    """
    Делает for_review попарно:
    status, Subclass_code, <db for Subclass_code>, Description, <db for Description>, ...
    затем несопоставленные db-колонки в конец.
    """
    df_compare = df_compare.copy()
    db_df = db_df.copy()

    if "Subclass_code" not in df_compare.columns:
        raise ValueError("В df_compare нет Subclass_code")
    if "Subclass_code" not in db_df.columns:
        raise ValueError("В db_df нет Subclass_code")

    merged = df_compare.merge(db_df, on="Subclass_code", how="left", suffixes=("", "_db"))

    export_cols: list[str] = []
    used_db_cols: set[str] = set()

    def add(col: str):
        if col in merged.columns and col not in export_cols:
            export_cols.append(col)

    # 0) status
    add("status")

    # 1) Subclass_code
    add("Subclass_code")

    # 2) пара для Subclass_code (если выбрана)
    target_db_code = (db_map or {}).get("Subclass_code")
    if target_db_code:
        add(target_db_code)
        used_db_cols.add(target_db_code)

    # 3) остальное по порядку UI (попарно)
    for src_col in cols_order:
        if src_col in ("Subclass_code", "status"):
            continue

        add(src_col)

        target_db = (db_map or {}).get(src_col)
        if target_db:
            add(target_db)
            used_db_cols.add(target_db)

    # 4) оставшиеся DB-колонки в конец (если нужны)
    for c in db_df.columns:
        if c == "Subclass_code":
            continue
        if c not in used_db_cols:
            add(c)

    return merged[export_cols]


def write_excel_with_highlight(
    buf: io.BytesIO,
    export_df: pd.DataFrame,
    highlight_cols: list[str],
    df_sections: pd.DataFrame | None = None,
    df_divisions: pd.DataFrame | None = None,
    df_groups_cmp: pd.DataFrame | None = None,
    df_classes_cmp: pd.DataFrame | None = None,
    df_subclasses_cmp: pd.DataFrame | None = None,
    debug: bool = False,
):
    """Подсвечивает только SHAMS-колонки (highlight_cols) на листе for_review."""
    fill = PatternFill(fill_type="solid", start_color="FFFFF2CC", end_color="FFFFF2CC")

    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        export_df.to_excel(writer, index=False, sheet_name="for_review")
        ws = writer.sheets["for_review"]

        col_to_idx = { _norm_col(name): i + 1 for i, name in enumerate(export_df.columns) }

        highlight_idxs = []
        for c in highlight_cols:
            idx = col_to_idx.get(_norm_col(c))
            if idx is not None:
                highlight_idxs.append(idx)

        if debug:
            print("EXPORT COLS:", list(export_df.columns))
            print("HIGHLIGHT COLS:", highlight_cols)
            print("HIGHLIGHT IDXS:", highlight_idxs)

        max_row = ws.max_row
        for col_idx in highlight_idxs:
            for row_idx in range(1, max_row + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

        # Дополнительные листы
        if df_sections is not None:
            df_sections.to_excel(writer, index=False, sheet_name="sections")
        if df_divisions is not None:
            df_divisions.to_excel(writer, index=False, sheet_name="divisions")

        # ВАЖНО: эти листы — именно сравнение
        if df_groups_cmp is not None:
            df_groups_cmp.to_excel(writer, index=False, sheet_name="groups")
        if df_classes_cmp is not None:
            df_classes_cmp.to_excel(writer, index=False, sheet_name="classes")
        if df_subclasses_cmp is not None:
            df_subclasses_cmp.to_excel(writer, index=False, sheet_name="subclasses")


# ================== UI ==================
st.title("Список активити провайдера")
st.markdown("---")


# ==================================================
# =============== STAGE 1 — UPLOAD =================
# ==================================================


if st.session_state.stage == STAGE_UPLOAD:
    st.markdown("### Выберите провайдера (здесь парсер подтянет provider_id из БД)")
    provider = st.selectbox(
        label="",
        options=["SHAMS", "Meydan", "IFZA"],
        index=["SHAMS", "Meydan", "IFZA"].index(st.session_state.get("selected_provider", "SHAMS")),
        key="selected_provider",
    )

    st.markdown("---")

    st.subheader("Укажите новый источник")

    uploaded = st.file_uploader("Загрузите файл shams2", type=["xlsx"])
    if uploaded is not None:
        st.session_state.shams2_bytes = uploaded.read()

    c1, c2 = st.columns(2)
    with c1:
        if st.button("Отменить"):
            st.session_state.shams2_bytes = None

    with c2:
        if st.button("Применить", disabled=st.session_state.shams2_bytes is None):
            load_shams()

            h_old, h_new, _ = build_header_change_log_from_bytes(
                st.session_state.shams_bytes,
                st.session_state.shams2_bytes,
                sheets=None,
            )

            st.session_state.headers_old = h_old
            st.session_state.headers_new = h_new
            st.session_state.headers_new_selected = list(h_new)

            # сбросы
            st.session_state.column_mapping = None
            st.session_state.hier_col_roles = None
            st.session_state.df_compare = None
            st.session_state.compare_stats = None
            st.session_state.db_column_mapping = None
            st.session_state.db_mapping_saved = False
            st.session_state.db_cols_order = None
            st.session_state.df_groups_cmp = None
            st.session_state.df_classes_cmp = None
            st.session_state.df_subclasses_cmp = None

            st.session_state.stage = STAGE_SELECT_HEADERS
            st.rerun()


# ==================================================
# =========== STAGE 2 — SELECT HEADERS =============
# ==================================================
if st.session_state.stage == STAGE_SELECT_HEADERS:
    st.subheader("Шаг 1 — выбор столбцов нового файла (shams2)")
    st.caption("Отметьте столбцы, которые пойдут в сопоставление")

    headers = st.session_state.headers_new or []
    prev_selected = st.session_state.headers_new_selected or []

    temp_selected = []
    for col in headers:
        checked = st.checkbox(col, value=(col in prev_selected), key=f"chk_{col}")
        if checked:
            temp_selected.append(col)

    st.session_state.headers_new_selected = temp_selected

    st.markdown("---")
    c1, c2 = st.columns(2)
    with c1:
        if st.button("Назад"):
            st.session_state.stage = STAGE_UPLOAD
            st.rerun()
    with c2:
        if st.button("Перейти к сопоставлению", disabled=len(temp_selected) == 0):
            st.session_state.column_mapping = None
            st.session_state.hier_col_roles = None
            st.session_state.df_compare = None
            st.session_state.compare_stats = None
            st.session_state.db_column_mapping = None
            st.session_state.db_mapping_saved = False
            st.session_state.db_cols_order = None
            st.session_state.df_groups_cmp = None
            st.session_state.df_classes_cmp = None
            st.session_state.df_subclasses_cmp = None

            st.session_state.stage = STAGE_MAPPING
            st.rerun()


# ==================================================
# ============== STAGE 3 — MAPPING =================
# ==================================================
if st.session_state.stage == STAGE_MAPPING:
    st.subheader("Шаг 2 — ручное сопоставление столбцов")
    st.caption(
        "Для каждого выбранного столбца из НОВОГО файла выберите соответствующий столбец "
        "в СТАРОМ файле. Если соответствия нет — оставьте «<нет соответствия>»."
    )

    headers_old = st.session_state.headers_old or []
    headers_new_selected = st.session_state.headers_new_selected or []

    if st.session_state.column_mapping is None:
        st.session_state.column_mapping = {col: None for col in headers_new_selected}
    else:
        current = {k: v for k, v in st.session_state.column_mapping.items() if k in headers_new_selected}
        for col in headers_new_selected:
            current.setdefault(col, None)
        st.session_state.column_mapping = current

    mapping = st.session_state.column_mapping

    st.markdown("---")
    for col_new in headers_new_selected:
        st.markdown(f"**{col_new} →**")
        options = ["<нет соответствия>"] + headers_old
        current_value = mapping.get(col_new)

        index = (headers_old.index(current_value) + 1) if (current_value in headers_old) else 0
        selected = st.selectbox(
            f"Соответствие для {col_new}",
            options=options,
            index=index,
            key=f"map_{col_new}",
        )
        mapping[col_new] = None if selected == "<нет соответствия>" else selected

    st.session_state.column_mapping = mapping

    st.markdown("---")
    c1, c2 = st.columns(2)
    with c1:
        if st.button("Назад"):
            st.session_state.stage = STAGE_SELECT_HEADERS
            st.rerun()
    with c2:
        if st.button("Далее: уровни иерархии"):
            st.session_state.hier_col_roles = None
            st.session_state.df_compare = None
            st.session_state.compare_stats = None
            st.session_state.db_column_mapping = None
            st.session_state.db_mapping_saved = False
            st.session_state.db_cols_order = None
            st.session_state.df_groups_cmp = None
            st.session_state.df_classes_cmp = None
            st.session_state.df_subclasses_cmp = None
            st.session_state.provider_has_groups = False
            st.session_state.provider_has_classes = False

            st.session_state.stage = STAGE_HIERARCHY
            st.rerun()


# ==================================================
# ============ STAGE 3.5 — HIERARCHY ===============
# ==================================================
# ==================================================
# ============ STAGE 3.5 — HIERARCHY ===============
# ==================================================
if st.session_state.stage == STAGE_HIERARCHY:
    st.subheader('Шаг 3 — "Уровни иерархии"')
    st.caption("Укажите, как провайдер структурирует активити, и назначьте роль каждому выбранному столбцу.")

    # --- ВАЖНО: добавь эти ключи в init_state() defaults ---
    # "provider_has_groups": False,
    # "provider_has_classes": False,

    # --- чек-боксы (дословно) ---
    has_groups = st.checkbox(
        "Провайдер разделяет активити на группы",
        value=st.session_state.get("provider_has_groups", False),
        key="chk_provider_groups",
    )
    has_classes = st.checkbox(
        "Провайдер разделяет активити на классы",
        value=st.session_state.get("provider_has_classes", False),
        key="chk_provider_classes",
    )

    # если есть группы — классы подразумеваются
    if has_groups and not has_classes:
        has_classes = True

    st.session_state.provider_has_groups = has_groups
    st.session_state.provider_has_classes = has_classes

    # --- варианты ролей в выпадающем списке ---
    # Никакого "не включать" (по твоей просьбе убрали)
    role_options: list[str] = ["Общий столбец", "Activity Code"]
    if has_classes:
        role_options.insert(1, "Class")
    if has_groups:
        role_options.insert(1, "Group")

    # --- кандидаты: выбранные колонки нового файла ---
    cols = st.session_state.headers_new_selected or []
    if not cols:
        st.warning("Нет выбранных колонок. Вернитесь назад.")
        if st.button("Назад"):
            st.session_state.stage = STAGE_MAPPING
            st.rerun()
        st.stop()

    # --- init roles (храним в st.session_state.hier_col_roles как раньше) ---
    if st.session_state.hier_col_roles is None:
        st.session_state.hier_col_roles = {c: "Общий столбец" for c in cols}

        # простые авто-эвристики
        for c in cols:
            cl = _norm_col(c)
            if cl == "subclass":
                st.session_state.hier_col_roles[c] = "Activity Code"
            elif cl == "class":
                st.session_state.hier_col_roles[c] = "Class" if has_classes else "Activity Code"
            elif cl == "group":
                st.session_state.hier_col_roles[c] = "Group" if has_groups else ("Class" if has_classes else "Activity Code")
            elif "description" in cl or cl in ("subclass_en", "desc"):
                st.session_state.hier_col_roles[c] = "Общий столбец"

    role_map = st.session_state.hier_col_roles

    # --- UI: один столбец ---
    for c in cols:
        cur = role_map.get(c, "Общий столбец")
        if cur not in role_options:
            cur = "Общий столбец"

        sel = st.selectbox(
            label=c,
            options=role_options,
            index=role_options.index(cur),
            key=f"hier_{c}",
        )
        role_map[c] = sel

    st.session_state.hier_col_roles = role_map

    st.markdown("---")
    c1, c2 = st.columns(2)
    with c1:
        if st.button("Назад"):
            st.session_state.stage = STAGE_MAPPING
            st.rerun()
    with c2:
        if st.button("Сохранить и перейти к статистике", type="primary"):
            st.session_state.df_compare = None
            st.session_state.compare_stats = None
            st.session_state.df_groups_cmp = None
            st.session_state.df_classes_cmp = None
            st.session_state.df_subclasses_cmp = None
            st.session_state.stage = STAGE_COMPARE
            st.rerun()



# ==================================================
# ============== STAGE 4 — COMPARE =================
# ==================================================
if st.session_state.stage == STAGE_COMPARE:
    st.subheader("Статистика сравнения")

    if st.session_state.df_compare is None:
        df_full_old, sec_old, div_old, grp_old, cls_old, sub_old = parse_all_sheets_from_bytes(
            st.session_state.shams_bytes, sheets=None
        )
        df_full_new, sec_new, div_new, grp_new, cls_new, sub_new = parse_all_sheets_from_bytes(
            st.session_state.shams2_bytes, sheets=None
        )

        # Основное сравнение активити (как раньше)
        df_compare = compare_shams(
            df_full_old,
            df_full_new,
            st.session_state.column_mapping,
        )
        st.session_state.df_compare = df_compare
        st.session_state.compare_stats = comparison_stats(df_compare)

        # Доп. листы сравнения уровней: groups/classes/subclasses (как Subclass)
        # Если у тебя на уровнях другие колонки — поменяй desc_col здесь.
        st.session_state.df_groups_cmp = compare_level_like_subclass(
            grp_old, grp_new, key_col="Group", desc_col="Group_en", out_key_name="Group"
        )
        st.session_state.df_classes_cmp = compare_level_like_subclass(
            cls_old, cls_new, key_col="Class", desc_col="Class_en", out_key_name="Class"
        )
        st.session_state.df_subclasses_cmp = compare_level_like_subclass(
            sub_old, sub_new, key_col="Subclass", desc_col="Subclass_en", out_key_name="Subclass"
        )

    stats_df = st.session_state.compare_stats
    stats = dict(zip(stats_df["metric"], stats_df["value"]))

    st.markdown(
        f"""
**Количество активити в старом файле:** {stats.get('Количество строк в старом файле', '')}  
**Количество активити в новом файле:** {stats.get('Количество строк в новом файле', '')}  
**Добавлено активити:** {stats.get('Добавлено', '')}  
**Удалено активити:** {stats.get('Удалено', '')}  
**Внесены изменения:** {stats.get('Изменено (по выбранным столбцам)', '')}  
**Остались без изменений:** {stats.get('Не изменено', '')}  
"""
    )

    c1, c2 = st.columns(2)
    with c1:
        if st.button("Назад"):
            st.session_state.stage = STAGE_HIERARCHY
            st.rerun()
    with c2:
        if st.button("Выгрузить в эксель для работы с обновлениями", type="primary"):
            st.session_state.stage = STAGE_DB_MAPPING
            st.rerun()


# ==================================================
# ============ STAGE 5 — DB MAPPING =================
# ==================================================
if st.session_state.stage == STAGE_DB_MAPPING:
    st.subheader("Сопоставление столбцов результата и Базы Данных")
    st.caption("Status не сопоставляется. Subclass_code сопоставляется отдельно и будет во второй позиции после status в выгрузке.")
    st.button("#### Тут будет кнопка выбора источника из БД (или автоматически подтягиваться таблица из БД provider_activity_type с provider_id = SHAMS id)")
    st.caption("Пока что автоматически используется файл shams_edit1.xlsx")

    df = st.session_state.df_compare
    if df is None or df.empty:
        st.error("Нет результата сравнения. Вернитесь на шаг сравнения.")
        st.stop()

    # если вдруг старый формат compare
    legacy_cols = [c for c in df.columns if c.endswith("_old") or c.endswith("_new") or c == "diff_columns"]
    if legacy_cols:
        st.warning("Найдены признаки старого df_compare (*_old/_new/diff_columns). Пересчитайте сравнение.")
        st.stop()

    cols_to_map: list[str] = []

    # 1) Subclass_code обязателен
    if "Subclass_code" in df.columns:
        cols_to_map.append("Subclass_code")

    # 2) всё остальное (кроме служебных)
    other = [c for c in df.columns if c not in ("Subclass_code", "status", "Subclass")]
    if "Description" in other:
        other = ["Description"] + [c for c in other if c != "Description"]

    cols_to_map += other
    cols_to_map = list(dict.fromkeys(cols_to_map))

    # запоминаем порядок UI — он нужен для попарного экспорта
    st.session_state.db_cols_order = cols_to_map

    # init mapping
    current_map = st.session_state.db_column_mapping or {}
    current_map = {k: v for k, v in current_map.items() if k in cols_to_map}
    for c in cols_to_map:
        current_map.setdefault(c, None)

    st.session_state.db_column_mapping = current_map
    mapping = st.session_state.db_column_mapping

    # один столбец UI
    for col in cols_to_map:
        cur_val = mapping.get(col)
        selected = st.selectbox(
            label=col,
            options=["<нет соответствия>"] + DB_COLUMNS,
            index=(DB_COLUMNS.index(cur_val) + 1) if cur_val in DB_COLUMNS else 0,
            key=f"db_map_{col}",
        )
        mapping[col] = None if selected == "<нет соответствия>" else selected

    st.session_state.db_column_mapping = mapping

    st.markdown("---")
    c1, c2 = st.columns(2)
    with c1:
        if st.button("Назад"):
            st.session_state.stage = STAGE_COMPARE
            st.rerun()
    with c2:
        if st.button("Сохранить сопоставление", type="primary"):
            st.session_state.db_mapping_saved = True
            st.session_state.stage = STAGE_DB_EXPORT
            st.rerun()


# ==================================================
# ============ STAGE 6 — DB EXPORT ==================
# ==================================================
if st.session_state.stage == STAGE_DB_EXPORT:
    st.subheader("Экспорт в Excel")

    if not st.session_state.get("db_mapping_saved"):
        st.warning("Сначала нажмите «Сохранить сопоставление».")
        st.stop()

    df_compare = st.session_state.df_compare
    if df_compare is None or df_compare.empty:
        st.error("Нет данных для экспорта. Вернитесь на шаг сравнения.")
        st.stop()

    db_map = st.session_state.db_column_mapping or {}
    cols_order = st.session_state.get("db_cols_order") or []

    # DB
    db_df = load_db_df()
    if db_df is None or db_df.empty:
        st.error("Файл БД пустой или не загрузился.")
        st.stop()

    # for_review (попарно)
    export_df = _build_export_df(
        df_compare=df_compare,
        db_df=db_df,
        db_map=db_map,
        cols_order=cols_order,
    )

    # уровни (сырые) из shams2 — оставляем как справочник
    try:
        _, df_sections, df_divisions, _, _, _ = parse_all_sheets_from_bytes(st.session_state.shams2_bytes, sheets=None)
    except Exception as e:
        st.error(f"Не удалось распарсить уровни из shams2: {e}")
        st.stop()

    # листы сравнения уровней (как Subclass)
    df_groups_cmp = st.session_state.get("df_groups_cmp")
    df_classes_cmp = st.session_state.get("df_classes_cmp")
    df_subclasses_cmp = st.session_state.get("df_subclasses_cmp")

    # подсветка только SHAMS-колонок на for_review
    # SHAMS = статус + все src колонки в UI-порядке (Subclass_code и дальше)
    highlight_cols = ["status"] + cols_order

    buf = io.BytesIO()
    write_excel_with_highlight(
        buf=buf,
        export_df=export_df,
        highlight_cols=highlight_cols,
        df_sections=df_sections,
        df_divisions=df_divisions,
        df_groups_cmp=df_groups_cmp,
        df_classes_cmp=df_classes_cmp,
        df_subclasses_cmp=df_subclasses_cmp,
        debug=False,  # поставь True на 1 запуск, если снова не окрасит
    )
    buf.seek(0)
    xlsx_bytes = buf.getvalue()

    st.download_button(
        label="Скачать в excel",
        data=xlsx_bytes,
        file_name="shams_compare_for_review.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.markdown("---")
    c1, c2 = st.columns(2)
    with c1:
        if st.button("Назад к сопоставлению с БД"):
            st.session_state.stage = STAGE_DB_MAPPING
            st.rerun()
    with c2:
        if st.button("Назад к статистике"):
            st.session_state.stage = STAGE_COMPARE
            st.rerun()

    st.button("#### Тут будет кнопка загрузки отредактированного менеджером файла")
    st.caption("Загруженный файл будет сопоставляться со столбцами в БД? или при загрузке будет автоматическое сопоставление?")

    # ===================== DRAFT UI: provider + pre-DB mapping =====================

    def build_ru_to_db_field_mapping() -> dict[str, str]:
        return {
            "ID": "id",
            "Официальное Наименование бизнес-деятельности у провайдера en": "title_by_provider_en",
            "Официальное Наименование бизнес-деятельности у провайдера ru": "title_by_provider_ru",
            "Введите код бизнес-деятельности": "business_activity_code",
            "Выберите группу": "activity_type_group_id",
            "Выберите класс": "activity_type_class_id",
            "ИД Универсальное наименование бизнес-деятельности": "business_activity_id",
            "Тип лицензии": "license_type_id",
            "Приоритет": "priority",
            "Описание вида деятельности en (=ПУСТО)": "description_en",
            "Описание вида деятельности ru (=ПУСТО)": "description_ru",
            "Примечание (для внутреннего использования)": "note",
            "Дополнительные требования, условия ": "note",
            "Нужны дополнительные разрешения (NOC)": "need_additional_permissions",
            "Существуют специальные требования к уставному капиталу": "requirements_for_authorized_capital",
            "Существуют специальные требования к инфраструктуре": "there_are_special_requirements_for_infrastructure",
            "Существуют специальные требования к учредителю": "there_are_special_requirements_for_founder",
            "Можно совмещать с другими активити": "can_be_combined_with_other_activities",
            "Деятельность только на территории страны регистрации": "activities_only_within_territory_of_country_of_registration",
            "Деятельность только за пределами страны регистрации": "activities_only_outside_country_of_registration",
            "Только для филиалов иностранных компаний": "for_branches_of_foreign_companies_only",
            "Существуют дополнительные требования, условия": "there_are_additional_requirements",
            "Пакеты": "packages",
            "Требуется наличие инфраструктурных объектов": "infrastructure_facilities_are_required",
            "Кто может быть учредителем": "who_can_be_founder",
            "Специальные требования к уставному капиталу (текст)": "authorized_capital_min_amount",
            "ОПФ": "possible_legal_form",
            "1. ИД органа": "additional_permission_1_authority_id",
            "1. ИД Услуги": "additional_permission_1_service_id",
            "2. ИД органа": "additional_permission_2_authority_id",
            "2. ИД Услуги": "additional_permission_2_service_id",
            "3. ИД органа": "additional_permission_3_authority_id",
            "3. ИД Услуги": "additional_permission_3_service_id",
        }


    def normalize_header_name(value: str) -> str:
        if value is None:
            return ""
        s = str(value).replace("\u00A0", " ")
        s = " ".join(s.split())
        return s.strip().lower()


    def auto_map_manager_columns_to_db(headers: list[str], mapping_dict: dict[str, str]) -> pd.DataFrame:
        norm_mapping = {normalize_header_name(k): v for k, v in mapping_dict.items()}

        rows = []
        for col in headers:
            db_field = norm_mapping.get(normalize_header_name(col))
            rows.append(
                {
                    "Столбец в файле менеджера": col,
                    "Поле в БД": db_field if db_field else "нет сопоставления с полями в БД",
                }
            )

        return pd.DataFrame(rows)


    # ===================== AUTO UI: manager file -> DB mapping =====================

    st.markdown("---")
    st.markdown("### Сопоставление столбцов файла менеджера с полями БД")

    uploaded_manager_file = st.file_uploader(
        "1) Загрузите xlsx, который обработал менеджер",
        type=["xlsx"],
        key="uploaded_manager_file_for_db_mapping",
    )

    if uploaded_manager_file is not None:
        st.session_state.manager_processed_bytes = uploaded_manager_file.read()

    c1, c2 = st.columns(2)

    with c1:
        if st.button("Очистить загруженный файл менеджера"):
            st.session_state.manager_processed_bytes = None
            st.session_state.manager_db_mapping_result = None
            st.rerun()

    with c2:
        if st.button(
                "2) Найти сопоставления с полями БД автоматически",
                disabled=st.session_state.manager_processed_bytes is None,
                type="primary",
        ):
            try:
                df_manager = pd.read_excel(io.BytesIO(st.session_state.manager_processed_bytes))
                manager_headers = list(df_manager.columns)

                ru_to_db_mapping = build_ru_to_db_field_mapping()
                mapping_result_df = auto_map_manager_columns_to_db(
                    headers=manager_headers,
                    mapping_dict=ru_to_db_mapping,
                )

                st.session_state.manager_db_mapping_result = mapping_result_df

            except Exception as e:
                st.session_state.manager_db_mapping_result = None
                st.error(f"Не удалось обработать файл менеджера: {e}")

    mapping_result_df = st.session_state.get("manager_db_mapping_result")

    if mapping_result_df is not None and not mapping_result_df.empty:
        st.markdown("### 3) Результат автоматического сопоставления")

        left_col, right_col = st.columns(2)

        with left_col:
            st.markdown("**Столбцы на русском**")
            for val in mapping_result_df["Столбец в файле менеджера"].tolist():
                st.write(val)

        with right_col:
            st.markdown("**Найденное поле в БД**")
            for val in mapping_result_df["Поле в БД"].tolist():
                st.write(val)

        st.markdown("---")
        st.dataframe(mapping_result_df, use_container_width=True)